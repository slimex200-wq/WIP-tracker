"""
WIP Deadline Checker - Windows 알림 스크립트
pip install plyer openpyxl schedule windows-toasts
"""

import os, json, time, threading, zipfile, re
import xml.etree.ElementTree as ET
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime, date, timedelta
from pathlib import Path

import schedule
import openpyxl
from plyer import notification

try:
    from windows_toasts import InteractableWindowsToaster, Toast, ToastDisplayImage, ToastImagePosition, ToastDuration
    _toaster = InteractableWindowsToaster("WIP Deadline Checker")
    _USE_WIN_TOAST = True
except Exception:
    _USE_WIN_TOAST = False

CONFIG_FILE  = Path.home() / ".wip_notifier_config.json"
IMAGE_CACHE  = Path.home() / ".wip_notifier_images"
ALERT_THRESHOLD = 14   # D-14 이하면 매일 알람

# ── 엑셀 헤더 이름 ──
COL_STYLE  = "Style Number"
COL_DESC   = "Description"
COL_FIT_DL = "Fit Approval"   # → "Fit Approval deadline"
COL_PPS_DL = "PPS Approval"   # → "PPS Approval Deadline"

# Fit Status 승인 키워드 (HTML 앱과 동일)
FIT_APPROVED_KW = [
    "approved", "approval", "c/o", "fixed", "complete", "done",
    "proceed to bulk", "proceed to pps", "submit pps",
    "go to pps", "to pps", "same silo",
]

ALERT_LABELS = {
    "block":  "🚫 PP 진행 불가",
    "urgent": "🔥 긴급",
}

def is_fit_approved(text):
    if not text: return False
    return any(k in str(text).lower() for k in FIT_APPROVED_KW)

def ddMD(d):
    """HTML 앱과 동일한 D-day 계산 (연도 무시, 월/일 기준)"""
    if not d: return None
    today = date.today()
    t = date(today.year, d.month, d.day)
    return (t - today).days

def classify_alert(fitOk, trimOk, fadDd, ppadDd):
    """HTML 앱과 동일한 4단계 경보 판단"""
    isUrgentFad  = not fitOk and fadDd is not None and 0 <= fadDd <= 14
    isUrgentPpad = ppadDd is not None and 0 <= ppadDd <= 14
    isTrimBlock  = not trimOk and ppadDd is not None and 0 <= ppadDd <= 14

    if not fitOk and fadDd is not None and fadDd < 0:       return "block"
    if isTrimBlock:                                          return "block"
    if isUrgentFad or isUrgentPpad:                          return "urgent"
    if not trimOk and ppadDd is not None and ppadDd <= 30:   return "trimwarn"
    if not fitOk and fadDd is not None and 0 <= fadDd <= 30: return "fitno"
    return None


# ── 설정 저장/불러오기 ──────────────────────────────────────

def load_config():
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"excel_path": "", "morning_time": "09:00", "notified_days": {}}

def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


# ── 날짜 파싱 ──────────────────────────────────────────────

def parse_date(val):
    if val is None: return None
    if isinstance(val, (datetime, date)):
        return val.date() if isinstance(val, datetime) else val
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m/%d", "%m-%d"):
            try:
                d = datetime.strptime(val.strip(), fmt)
                if d.year == 1900: d = d.replace(year=date.today().year)
                return d.date()
            except ValueError:
                continue
    return None

def find_col(headers, name):
    for i, h in enumerate(headers):
        if h and name.lower() in str(h).lower(): return i
    return None


# ── 엑셀 데이터 읽기 ────────────────────────────────────────

def read_wip_data(excel_path):
    """HTML 앱과 동일한 파싱 로직 — block/urgent 경보 스타일 반환"""
    if not excel_path or not os.path.exists(excel_path): return []
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"[오류] 엑셀 열기 실패: {e}"); return []

    items = []

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2: continue

        # Brand Moment 헤더 행 찾기 (HTML 앱과 동일)
        header_row, header_idx = None, 0
        for i, row in enumerate(rows[:10]):
            if not row: continue
            for c in range(3):
                if c < len(row) and row[c] and str(row[c]).strip().lower() == "brand moment":
                    header_row = [str(cell).strip() if cell else "" for cell in row]
                    header_idx = i; break
            if header_row: break
        if header_row is None: continue

        # 컬럼 매칭 (HTML 앱과 동일)
        fad_col  = find_col(header_row, "fit approval")
        fst_col  = find_col(header_row, "status")  # exact match
        if fst_col is not None and header_row[fst_col].lower() != "status":
            fst_col = None
        ppad_col = find_col(header_row, "pps approval")
        trim_col = find_col(header_row, "trim app")

        for row in rows[header_idx + 1:]:
            if not row or not any(row): continue
            def get(col): return row[col] if col is not None and col < len(row) else None

            style = str(get(1) or "").strip()  # 인덱스 1 고정
            desc  = str(get(2) or "").strip()  # 인덱스 2 고정
            if not style: continue

            fadDate  = parse_date(get(fad_col))
            ppadDate = parse_date(get(ppad_col))
            trimDate = parse_date(get(trim_col))
            fadSt    = str(get(fst_col) or "").strip() if fst_col else ""

            fadDd  = ddMD(fadDate)
            ppadDd = ddMD(ppadDate)

            fitOk  = is_fit_approved(fadSt)
            trimOk = False  # openpyxl read_only에서 gray shade 감지 불가 → 미완료로 간주

            alert = classify_alert(fitOk, trimOk, fadDd, ppadDd)

            if alert in ("block", "urgent"):
                # 가장 급한 날짜 선택
                urgent_date = None
                urgent_type = ""
                if not fitOk and fadDd is not None and fadDd <= 14:
                    urgent_date = fadDate; urgent_type = "FAD"
                if ppadDd is not None and ppadDd <= 14:
                    if urgent_date is None or (ppadDate and ppadDate < urgent_date):
                        urgent_date = ppadDate; urgent_type = "PPAD"

                items.append({
                    "style": style, "desc": desc,
                    "sheet": sheet.title, "alert": alert,
                    "type": urgent_type,
                    "date": urgent_date,
                    "diff": ddMD(urgent_date) if urgent_date else 0,
                })

    wb.close()
    return items


# ── 스케치 이미지 추출 ─────────────────────────────────────

def extract_style_images(excel_path):
    """xlsx Sketch 컬럼 이미지를 추출해 {style: 파일경로} 반환"""
    IMAGE_CACHE.mkdir(exist_ok=True)
    style_images = {}

    _NS = {
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'r':   'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
        'a':   'http://schemas.openxmlformats.org/drawingml/2006/main',
    }
    _EMBED = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed'

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        with zipfile.ZipFile(excel_path) as z:
            for sheet_idx, sheet_name in enumerate(sheet_names, start=1):
                # 시트 → drawing 파일 경로
                ws_rels = f'xl/worksheets/_rels/sheet{sheet_idx}.xml.rels'
                if ws_rels not in z.namelist(): continue
                drawing_path = None
                for r in ET.fromstring(z.read(ws_rels)):
                    if 'drawing' in r.get('Target', '').lower():
                        drawing_path = r.get('Target').replace('../', 'xl/')
                        break
                if not drawing_path or drawing_path not in z.namelist(): continue

                # drawing → rId→이미지 파일 매핑
                num = re.search(r'drawing(\d+)', drawing_path).group(1)
                drw_rels = f'xl/drawings/_rels/drawing{num}.xml.rels'
                if drw_rels not in z.namelist(): continue
                rid_map = {}
                for r in ET.fromstring(z.read(drw_rels)):
                    rid_map[r.get('Id')] = r.get('Target').replace('../', 'xl/')

                # Sketch 컬럼(col=3) anchor row → 이미지 파일
                anchor_images = []
                for anchor in ET.fromstring(z.read(drawing_path)):
                    tag = anchor.tag.split('}')[-1]
                    if tag not in ('twoCellAnchor', 'oneCellAnchor'): continue
                    from_el = anchor.find('xdr:from', _NS)
                    if from_el is None: continue
                    col_el = from_el.find('xdr:col', _NS)
                    row_el = from_el.find('xdr:row', _NS)
                    if col_el is None or row_el is None: continue
                    if int(col_el.text) != 3: continue   # Sketch 컬럼만
                    blip = anchor.find('.//a:blip', _NS)
                    if blip is None: continue
                    img_file = rid_map.get(blip.get(_EMBED), '')
                    if img_file.endswith('.png') and img_file in z.namelist():
                        anchor_images.append((int(row_el.text), img_file))

                if not anchor_images: continue

                # 시트 row → style 매핑
                wb2 = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
                rows = list(wb2[sheet_name].iter_rows(values_only=True))
                wb2.close()

                header_idx = 0
                for i, row in enumerate(rows[:5]):
                    rs = [str(c).upper() if c else '' for c in row]
                    if any('FIT APPROVAL' in c or 'PPS APPROVAL' in c for c in rs):
                        header_idx = i; break

                row_style = {}
                for ri, row in enumerate(rows[header_idx + 1:], start=header_idx + 1):
                    if row[1] and str(row[1]).strip() not in ('', 'None'):
                        row_style[ri] = str(row[1]).strip()

                sorted_rows = sorted(row_style.keys())

                # 이미지 → style 매핑 (style당 첫 번째 이미지만)
                for anchor_row, img_file in sorted(anchor_images):
                    matched = next((r for r in reversed(sorted_rows) if r <= anchor_row), None)
                    if matched is None: continue
                    style = row_style[matched]
                    if style in style_images: continue  # 이미 있으면 skip
                    safe = re.sub(r'[^\w\-]', '_', style)
                    img_path = IMAGE_CACHE / f"{safe}.png"
                    if not img_path.exists():
                        img_path.write_bytes(z.read(img_file))
                    style_images[style] = str(img_path)

    except Exception as e:
        print(f"[이미지 추출 오류] {e}")

    return style_images


# ── 알림 전송 ──────────────────────────────────────────────

def send_notification(title, message, image_path=None, timeout=8):
    if _USE_WIN_TOAST:
        try:
            toast = Toast(text_fields=[title, message], duration=ToastDuration.Long)
            if image_path and os.path.exists(image_path):
                toast.AddImage(ToastDisplayImage.fromPath(
                    image_path, position=ToastImagePosition.Hero))
            _toaster.show_toast(toast)
            return
        except Exception as e:
            print(f"[Win알림 오류] {e}")
    # fallback
    try:
        notification.notify(title=title, message=message,
                            app_name="WIP Deadline Checker", timeout=timeout)
    except Exception as e:
        print(f"[알림 오류] {e}")


# ── 아침 요약 ──────────────────────────────────────────────

def morning_summary(cfg):
    items = read_wip_data(cfg.get("excel_path", ""))
    today = date.today()

    blocks  = [i for i in items if i["alert"] == "block"]
    urgents = [i for i in items if i["alert"] == "urgent"]

    lines = [f"{today.strftime('%m/%d')} WIP 긴급 현황"]
    if blocks:  lines.append(f"🚫 PP 진행 불가: {len(blocks)}건")
    if urgents: lines.append(f"🔥 긴급: {len(urgents)}건")
    if not blocks and not urgents:
        lines.append("긴급 항목 없음 ✓")

    # 상위 3건 스타일 표시
    top = (blocks + urgents)[:3]
    for i in top:
        d = i["diff"]
        lines.append(f"  {i['style']} [{i['type']}] D{'+' if d < 0 else '-'}{abs(d)}")

    send_notification("WIP 긴급 현황 요약", "\n".join(lines))
    print(f"[{datetime.now().strftime('%H:%M')}] 아침 요약 전송 (block:{len(blocks)} urgent:{len(urgents)})")


# ── D-day 알림 체크 ────────────────────────────────────────

def check_deadline_alerts(cfg):
    excel_path = cfg.get("excel_path", "")
    items = read_wip_data(excel_path)
    if not items: return

    today_str = date.today().isoformat()
    notified  = cfg.get("notified_days", {})

    # 이미 오늘 알림 보냈으면 스킵
    if notified.get(f"_summary_{today_str}"): return

    blocks  = [i for i in items if i["alert"] == "block"]
    urgents = [i for i in items if i["alert"] == "urgent"]

    # 요약 1건으로 통합 알림
    lines = []
    if blocks:  lines.append(f"🚫 PP 진행 불가: {len(blocks)}건")
    if urgents: lines.append(f"🔥 긴급: {len(urgents)}건")

    # 가장 급한 상위 5건 표시
    top = sorted(blocks + urgents, key=lambda i: i["diff"] or 0)[:5]
    for i in top:
        d = i["diff"]
        dtxt = f"D{'+' if d < 0 else '-'}{abs(d)}" if d is not None else ""
        lines.append(f"  {i['style']} [{i['type']}] {dtxt}")

    if len(items) > 5:
        lines.append(f"  ...외 {len(items) - 5}건")

    send_notification(
        f"WIP 긴급 {len(blocks) + len(urgents)}건",
        "\n".join(lines)
    )
    notified[f"_summary_{today_str}"] = True
    print(f"[{datetime.now().strftime('%H:%M')}] 긴급 요약 알림 전송 (block:{len(blocks)} urgent:{len(urgents)})")

    cutoff = (date.today() - timedelta(days=30)).isoformat()
    cfg["notified_days"] = {k: v for k, v in notified.items() if k.split("_")[-1] >= cutoff}
    save_config(cfg)


# ── 스케줄러 ───────────────────────────────────────────────

def run_scheduler(cfg):
    morning_time = cfg.get("morning_time", "09:00")
    schedule.every().day.at(morning_time).do(morning_summary, cfg)
    schedule.every(1).hours.do(check_deadline_alerts, cfg)
    check_deadline_alerts(cfg)
    print(f"[WIP Notifier] 매일 {morning_time} 요약 + 1시간마다 D-day 체크")
    while True:
        schedule.run_pending()
        time.sleep(60)


# ── 설정 UI ────────────────────────────────────────────────

def open_settings(cfg, on_save=None):
    win = tk.Tk()
    win.title("WIP Notifier 설정")
    win.geometry("500x250")
    win.resizable(False, False)
    win.configure(bg="#1a1a1a")
    fg, bg, bg2, red = "#e8e8e8", "#1a1a1a", "#2a2a2a", "#f53d3d"

    def lbl(text, **kw):
        opts = {"bg": bg, "fg": fg, "font": ("Consolas", 9)}
        opts.update(kw)
        return tk.Label(win, text=text, **opts)

    lbl("엑셀 파일 경로").place(x=20, y=20)
    path_var = tk.StringVar(value=cfg.get("excel_path", ""))
    tk.Entry(win, textvariable=path_var, bg=bg2, fg=fg, insertbackground=fg,
             relief="flat", font=("Consolas", 9), bd=4).place(x=20, y=42, width=380, height=28)

    def browse():
        p = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel", "*.xlsx *.xlsm *.xls"), ("All", "*.*")])
        if p: path_var.set(p)
    tk.Button(win, text="찾기", command=browse, bg="#333", fg=fg,
              relief="flat", font=("Consolas", 9), cursor="hand2").place(x=408, y=42, height=28)

    lbl("아침 요약 알림 시간 (HH:MM)").place(x=20, y=90)
    time_var = tk.StringVar(value=cfg.get("morning_time", "09:00"))
    tk.Entry(win, textvariable=time_var, bg=bg2, fg=fg, insertbackground=fg,
             relief="flat", font=("Consolas", 9), bd=4).place(x=20, y=112, width=120, height=28)

    def save():
        path = path_var.get().strip()
        t    = time_var.get().strip()
        if not path:
            messagebox.showerror("오류", "엑셀 파일 경로를 입력해주세요."); return
        if not os.path.exists(path):
            messagebox.showerror("오류", f"파일을 찾을 수 없습니다:\n{path}"); return
        try: datetime.strptime(t, "%H:%M")
        except ValueError:
            messagebox.showerror("오류", "시간 형식이 올바르지 않습니다. (예: 09:00)"); return
        cfg["excel_path"]   = path
        cfg["morning_time"] = t
        save_config(cfg)
        messagebox.showinfo("저장 완료", "설정이 저장되었습니다!\n알림이 시작됩니다.")
        win.destroy()
        if on_save: on_save()

    tk.Button(win, text="저장하고 시작", command=save, bg=red, fg="white",
              relief="flat", font=("Consolas", 10, "bold"),
              padx=20, pady=6, cursor="hand2").place(x=20, y=170)
    lbl("창을 닫아도 백그라운드에서 계속 실행됩니다.", fg="#555").place(x=20, y=215)
    win.mainloop()


# ── 진입점 ─────────────────────────────────────────────────

def main():
    cfg = load_config()

    if not cfg.get("excel_path") or not os.path.exists(cfg["excel_path"]):
        saved = threading.Event()
        open_settings(cfg, on_save=lambda: saved.set())
        saved.wait(timeout=60)
        cfg = load_config()

    if not cfg.get("excel_path"):
        print("설정이 완료되지 않았습니다."); return

    t = threading.Thread(target=run_scheduler, args=(cfg,), daemon=True)
    t.start()
    print(f"[WIP Notifier] 실행 중 | 경로: {cfg['excel_path']} | 아침: {cfg['morning_time']}")

    try:
        while True: time.sleep(1)
    except KeyboardInterrupt:
        print("종료합니다.")

if __name__ == "__main__":
    main()
